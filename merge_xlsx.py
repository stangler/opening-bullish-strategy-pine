import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

XLSX_DIR = Path("xlsx")
OUTPUT_FILE = XLSX_DIR / "merged.xlsx"

# Files to merge: (filename, label)
files = [
    ("前場1分足.xlsx", "前場1分足"),
    ("前場3分足.xlsx", "前場3分足"),
    ("前場5分足.xlsx", "前場5分足"),
    ("後場1分足.xlsx", "後場1分足"),
    ("後場3分足.xlsx", "後場3分足"),
    ("後場5分足.xlsx", "後場5分足"),
]

# Read all data
all_rows = []
header = None
sheet_name = None

for filename, label in files:
    wb = openpyxl.load_workbook(XLSX_DIR / filename)
    ws = wb.active
    if sheet_name is None:
        sheet_name = ws.title
    
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    file_header = rows[0]
    
    if header is None:
        header = list(file_header)
    
    data_rows = rows[1:]
    for row in data_rows:
        all_rows.append((label,) + row)
    wb.close()

# Build new header with 時間足 column
new_header = ["時間足"] + header

# Create workbook
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = sheet_name or "merged"

# Styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Write header
for col_idx, val in enumerate(new_header, 1):
    cell = ws_out.cell(row=1, column=col_idx, value=val)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Write data
for row_idx, row in enumerate(all_rows, 2):
    for col_idx, val in enumerate(row, 1):
        cell = ws_out.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-adjust column widths
for col_idx in range(1, len(new_header) + 1):
    max_len = len(str(new_header[col_idx - 1]))
    for row_idx in range(2, len(all_rows) + 2):
        val = ws_out.cell(row=row_idx, column=col_idx).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
    ws_out.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

# Freeze top row
ws_out.freeze_panes = "A2"

# Auto-filter
ws_out.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(new_header))}{len(all_rows) + 1}"

wb_out.save(OUTPUT_FILE)
print(f"Merged {len(all_rows)} rows into {OUTPUT_FILE}")
print(f"Columns: {len(new_header)}")

# Quick summary
for label in [f[1] for f in files]:
    count = sum(1 for r in all_rows if r[0] == label)
    print(f"  {label}: {count} rows")