import csv
from pathlib import Path
from openpyxl import load_workbook

STRATEGY_KEYS = [
    "総損益", "最大ドローダウン",
    "勝率", "プロフィットファクター",
]

DW_KEYS = [
    "AM GapUp 陽", "AM GapUp 陰", "AM GapUp 勝率%", "AM GapUp EV",
    "AM GapDown 陽", "AM GapDown 陰", "AM GapDown 勝率%", "AM GapDown EV",
    "AM Cont 陽", "AM Cont 陰", "AM Cont 勝率%", "AM Cont EV",
    "AM Sum 陽", "AM Sum 陰", "AM Sum 勝率%", "AM Sum EV",
    "PM GapUp 陽", "PM GapUp 陰", "PM GapUp 勝率%", "PM GapUp EV",
    "PM GapDown 陽", "PM GapDown 陰", "PM GapDown 勝率%", "PM GapDown EV",
    "PM Cont 陽", "PM Cont 陰", "PM Cont 勝率%", "PM Cont EV",
    "PM Sum 陽", "PM Sum 陰", "PM Sum 勝率%", "PM Sum EV",
]

def extract_from_csv(csv_file):
    data = {}
    section = None  # "strategy" | "dw"

    with open(csv_file, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue

            # 銘柄行
            if row[0].strip() == "銘柄" and len(row) >= 2:
                data["銘柄"] = row[1].strip()
                continue

            # セクション切替
            if row[0].strip() == "=== Strategy Tester ===":
                section = "strategy"
                continue
            if row[0].strip() == "=== Data Window ===":
                section = "dw"
                continue

            # ヘッダー行スキップ
            if row[0].strip() == "指標":
                continue

            if len(row) < 2:
                continue

            key   = row[0].strip()
            value = row[1].strip()

            if section == "strategy" and key in STRATEGY_KEYS and key not in data:
                data[key] = value
            elif section == "dw" and key in DW_KEYS and key not in data:
                data[key] = value

    # トレード総数・勝ちトレード・負けトレードをDW値から計算
    def _int(key):
        v = data.get(key, "")
        try:
            # float経由でint化（"18.0"等の小数文字列に対応）
            return int(float(str(v).replace(",", "").replace("∅", "0")))
        except ValueError:
            return 0

    win  = _int("AM Sum 陽") + _int("PM Sum 陽")
    lose = _int("AM Sum 陰") + _int("PM Sum 陰")
    data["勝ちトレード"] = str(win)
    data["負けトレード"] = str(lose)
    data["トレード総数"] = str(win + lose)

    return data

def update_excel(csv_data_list, xlsx_path="format.xlsx"):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # ヘッダー → 列インデックス マッピング（0始まり）
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}

    all_keys = STRATEGY_KEYS + ["トレード総数", "勝ちトレード", "負けトレード"] + DW_KEYS

    for row in ws.iter_rows(min_row=2):
        symbol_cell = row[col_map["銘柄"]]
        if not symbol_cell.value:
            continue
        symbol = str(symbol_cell.value).strip()

        csv_data = next((d for d in csv_data_list if d.get("銘柄") == symbol), None)
        if not csv_data:
            print(f"  ✗ CSVなし: {symbol}")
            continue

        print(f"  ✓ {symbol}")
        for key in all_keys:
            if key in col_map and key in csv_data:
                row[col_map[key]].value = csv_data[key]

    wb.save(xlsx_path)
    print(f"\n保存 → {xlsx_path}")

def main():
    csv_files = list(Path(".").glob("*.csv"))
    print(f"CSV {len(csv_files)} 件\n")

    csv_data_list = []
    for f in csv_files:
        d = extract_from_csv(f)
        csv_data_list.append(d)
        print(f"{f.name}: 銘柄={d.get('銘柄')} / {len(d)-1}項目")

    print()
    update_excel(csv_data_list)

if __name__ == "__main__":
    main()